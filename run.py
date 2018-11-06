import tensorflow as tf
import numpy as np
import openpyxl
import os


tf.set_random_seed(777)

# train Parameters
seq_length = 7
data_dim = 5
hidden_dim = 5
output_dim = 1
num_stacked_layers = 3
learning_rate = 0.001
num_epoch = 5000
check_step = 10000

def MinMaxScaler(data):
    numerator = data - np.min(data, 0)
    denominator = np.max(data, 0) - np.min(data, 0)
    return numerator / (denominator+ 1e-7)

def reverse(data, origin):
    min = np.min(origin, 0)
    max = np.max(origin, 0)
    return data * (max[6] - min[6] + 1e-7) + min[6]

def read(s, e):
    dataX = []
    dataY = []
    for num in range(s, e+1):
        xy = np.load("./data/" + str(num) + ".npy")
        xy = MinMaxScaler(xy)
        x = xy[:, [0, 3, 4, 5, 6]]
        y = xy[:, 6]
        for idx in range(0, len(y) - seq_length):
            _x = x[idx: idx + seq_length]
            _y = y[idx + seq_length]
            dataX.append(_x)
            dataY.append(_y)
    return dataX, dataY

def lstm_cell(ReLu = False):
    if ReLu:
        return tf.contrib.rnn.BasicLSTMCell(num_units=hidden_dim, state_is_tuple=True, activation=tf.nn.relu)
    return tf.contrib.rnn.BasicLSTMCell(num_units=hidden_dim, state_is_tuple=True, activation=tf.tanh)


def rnn_cell():
    return tf.contrib.rnn.BasicRNNCell(num_units=hidden_dim, activation=tf.nn.relu)


def gru_cell(ReLu = False):
    if ReLu:
        return tf.contrib.rnn.GRUCell(num_units=hidden_dim, activation=tf.nn.relu)
    return tf.contrib.rnn.GRUCell(num_units=hidden_dim, activation=tf.nn.tanh)


trainData = read(2014, 2017)
testData = read(2018, 2018)

trainX, trainY = trainData[0], np.reshape(trainData[1], (-1, 1))
testX, testY = testData[0], np.reshape(testData[1], (-1, 1))


X = tf.placeholder(tf.float32, [None, seq_length, data_dim])
Y = tf.placeholder(tf.float32, [None, 1])

multi_cells = tf.contrib.rnn.MultiRNNCell([gru_cell(True) for _ in range(num_stacked_layers)], state_is_tuple=True)
outputs, _ = tf.nn.dynamic_rnn(multi_cells, X, dtype=tf.float32)
Y_pred = tf.contrib.layers.fully_connected(outputs[:, -1], output_dim, activation_fn=None)

loss = tf.reduce_sum(tf.square(Y_pred - Y))

optimizer = tf.train.AdamOptimizer(learning_rate)
train = optimizer.minimize(loss)

targets = tf.placeholder(tf.float32, [None, 1])
predictions = tf.placeholder(tf.float32, [None, 1])
rmse = tf.sqrt(tf.reduce_mean(tf.square(targets - predictions)))


with tf.Session() as sess:
    sess.run(tf.global_variables_initializer())

    for epoch in range(1, num_epoch + 1):
        _, step_loss = sess.run([train, loss], feed_dict = {X: trainX, Y: trainY})

        if epoch % check_step ==0:
            print("[step: {}] loss: {}".format(epoch, step_loss))

    test_predict = sess.run(Y_pred, feed_dict = {X: testX})
    rmse_val = sess.run(rmse, feed_dict={targets: testY, predictions: test_predict})
    print("RMSE: {}".format(rmse_val))


    test_xy = np.load("./data/2018.npy")
    test_y = reverse(testY, test_xy)

    error_sum = 0.
    max_error = 0.
    result = []
    for i in range(len(testX)):
        testX_one = np.reshape(testX[i], (-1, seq_length, data_dim))
        test_y_pred = sess.run(Y_pred, feed_dict={X: testX_one})
        test_y_pred = reverse(test_y_pred, test_xy)
        aa = abs(test_y[i, 0] - test_y_pred[0, 0]) / test_y[i, 0] * 100
        error_sum = error_sum + aa
        if max_error < aa:
            max_error = (test_y[i, 0] - test_y_pred[0, 0]) / test_y[i, 0] * 100
        result.append(test_y_pred[0, 0])

    print("average error: ", error_sum / len(testX), "%")
    print("max diff: ", max_error, "%")

    if not "score.xlsx" in os.listdir("./"):
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save("./score.xlsx")
        wb.close()

    wb = openpyxl.load_workbook('score.xlsx')
    ws = wb.active

    ws.cell(row=1, column=1).value = "Predict"
    ws.cell(row=1, column=2).value = "Truth"
    for i in range(1, len(testX) + 1):
        ws.cell(row=i + 1, column=1).value = result[i - 1]
        ws.cell(row=i + 1, column=2).value = test_y[i - 1, 0]

    wb.save("./score.xlsx")
    wb.close()
